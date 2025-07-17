import os
import shutil
import json
import pandas as pd
from openpyxl import Workbook
from Prep_and_Tools.io_tools import list_folder_contents
from Prep_and_Tools.patient_tools import load_case_config
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def create_results_structure(output_root, patient, case_type):
    """
    Handles smart creation of Results folder for one patient with dynamic visit handling.
    
    Returns:
        str: patient folder path
        bool: whether to skip the patient
    """
    patient_id = patient["patient_number"]
    patient_folder = os.path.join(output_root, "Results", case_type.upper(), f"PATIENT_{patient_id}")

    if os.path.exists(patient_folder):
        print(f"\n📁 Patient folder already exists: {patient_folder}")

        # Dynamically list contents of each visit folder
        list_folder_contents(f"Patient {patient_id}", patient_folder)
        for visit in patient["visits"]:
            visit_path = os.path.join(patient_folder, visit)
            list_folder_contents(f"{visit}", visit_path)

        decision = input("❓ Use existing results? [S]kip patient / [O]verwrite and reprocess: ").strip().lower()
        if decision == 's':
            print(f"⏩ Skipping patient {patient_id}")
            return patient_folder, True
        elif decision == 'o':
            print(f"🧨 Overwriting patient folder: {patient_folder}")
            shutil.rmtree(patient_folder)

    # Recreate visit folders (new or just cleared)
    os.makedirs(patient_folder, exist_ok=True)
    for visit in patient["visits"]:
        os.makedirs(os.path.join(patient_folder, visit), exist_ok=True)

    return patient_folder, False


def export_patient_list_to_excel(txt_paths, output_path):
    """
    Creates a Patient_List.xlsx file showing progress for each step.
    - Completed steps show a green background and a checkmark (✔️)
    - Missing steps show an empty cell
    """

    def _safe_exists(p):
        return isinstance(p, (str, bytes, os.PathLike)) and os.path.exists(p)

    display_rows = []
    color_flags_list = []

    for path in txt_paths:
        try:
            config = load_case_config(path)
            pid = config["patient_number"]
            case_type = config["case_type"]
            levels = config.get("level_intervened", [])
            levels_str = ", ".join(levels) if isinstance(levels, list) else str(levels)
            output_dir = config["patient_output_path"]

            # Default: everything False
            progress = {
                "PRE_NIfTI": False,
                "POST_NIfTI": False,
                "PRE_Seg": False,
                "POST_Seg": False,
                "PRE_Mesh": False,
                "POST_Mesh": False,
                "Registration": False,
                "Crop": False
            }

            # Look for JSON state
            json_path = os.path.join(output_dir, f"PATIENT_{pid}_state.json")
            if os.path.exists(json_path):
                with open(json_path, "r") as f:
                    data = json.load(f)

                pre = data.get("pre", {})
                post = data.get("post", {})

                progress["PRE_NIfTI"] = _safe_exists(pre.get("nifti_path"))
                progress["POST_NIfTI"] = _safe_exists(post.get("nifti_path"))

                seg_pre = pre.get("segmentation_paths", [])
                seg_post = post.get("segmentation_paths", [])
                mesh_pre = pre.get("mesh_paths", [])
                mesh_post = post.get("mesh_paths", [])
                reg = data.get("registration_paths", [])
                crop = data.get("crop_paths", [])

                progress["PRE_Seg"] = bool(seg_pre) and all(_safe_exists(p) for p in seg_pre)
                progress["POST_Seg"] = bool(seg_post) and all(_safe_exists(p) for p in seg_post)
                progress["PRE_Mesh"] = bool(mesh_pre) and all(_safe_exists(p) for p in mesh_pre)
                progress["POST_Mesh"] = bool(mesh_post) and all(_safe_exists(p) for p in mesh_post)
                progress["Registration"] = bool(reg) and all(_safe_exists(p) for p in reg)
                progress["Crop"] = bool(crop) and any(_safe_exists(p) for p in crop)

            else:
                progress["PRE_NIfTI"] = _safe_exists(config.get("pre_nifti_path"))
                progress["POST_NIfTI"] = _safe_exists(config.get("post_nifti_path"))

            # Display row with checkmarks or empty cells
            display_row = {
                "Patient_ID": pid,
                "Case_Type": case_type,
                "Levels_Intervened": levels_str,
                "PRE_NIfTI": "✔️" if progress["PRE_NIfTI"] else "",
                "POST_NIfTI": "✔️" if progress["POST_NIfTI"] else "",
                "PRE_Seg": "✔️" if progress["PRE_Seg"] else "",
                "POST_Seg": "✔️" if progress["POST_Seg"] else "",
                "PRE_Mesh": "✔️" if progress["PRE_Mesh"] else "",
                "POST_Mesh": "✔️" if progress["POST_Mesh"] else "",
                "Registration": "✔️" if progress["Registration"] else "",
                "Crop": "✔️" if progress["Crop"] else ""
            }

            display_rows.append(display_row)
            color_flags_list.append(progress)

        except Exception as e:
            print(f"⚠️ Could not process patient at {path}: {e}")
            continue

    if not display_rows:
        print("⚠️ No valid patients to export.")
        return None

    # Build data and styling frames
    df = pd.DataFrame(display_rows)
    excel_path = os.path.join(output_path, "Patient_Progression_List.xlsx")

    progress_cols = [
        "PRE_NIfTI", "POST_NIfTI",
        "PRE_Seg", "POST_Seg",
        "PRE_Mesh", "POST_Mesh",
        "Registration", "Crop"
    ]
    style_mask = pd.DataFrame(color_flags_list)[progress_cols]

    def apply_color(col):
        return [
            "background-color: lightgreen" if style_mask.loc[idx, col.name] else ""
            for idx in col.index
        ]

    styled = df.style.apply(apply_color, subset=progress_cols)
    styled.to_excel(excel_path, index=False)

    print(f"📄 Exported patient list to: {excel_path}")
    return excel_path


def create_analysis_results_excel(output_dir):
    """
    Creates an empty Excel file for future analysis results.
    Prompts the user for filename and whether to overwrite if it already exists.
    """
    print("\n🧾 Create a new (empty) Analysis Excel")

    filename = input("🆕 Enter Excel filename (default: 'Analysis_Results.xlsx'): ").strip()
    if not filename:
        filename = "Analysis_Results.xlsx"
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    excel_path = os.path.join(output_dir, filename)

    if os.path.exists(excel_path):
        print(f"⚠️ File already exists: {excel_path}")
        decision = input("Overwrite it? [Y/N]: ").strip().lower()
        if decision != "y":
            new_name = input("Enter a different filename: ").strip()
            if not new_name.endswith(".xlsx"):
                new_name += ".xlsx"
            excel_path = os.path.join(output_dir, new_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Volume Analysis"
    wb.save(excel_path)

    print(f"✅ Created empty Excel file at: {excel_path}")
    return excel_path


def save_spinal_cord_global_excel(areas_data, perims_data, excel_path, case1, case2):

    workbook = load_workbook(excel_path)

    # 📘 Create or access the 'Global Comparison' sheet
    if "Global Comparison" in workbook.sheetnames:
        sheet = workbook["Global Comparison"]
    else:
        sheet = workbook.create_sheet("Global Comparison")
        headers = ["Metric", f"{case1} Mean", f"{case1} STD", f"{case2} Mean", f"{case2} STD"]
        sheet.append(headers)
        for idx in range(1, 6):
            sheet.column_dimensions[get_column_letter(idx)].width = 20
        for cell in sheet[1]:
            cell.font = Font(bold=True)

    # Data rows
    sheet.append([
        "Areas",
        round(areas_data[case1]["global_mean"], 2),
        round(areas_data[case1]["global_std"], 2),
        round(areas_data[case2]["global_mean"], 2),
        round(areas_data[case2]["global_std"], 2)
    ])
    sheet.append([
        "Perimeters",
        round(perims_data[case1]["global_mean"], 2),
        round(perims_data[case1]["global_std"], 2),
        round(perims_data[case2]["global_mean"], 2),
        round(perims_data[case2]["global_std"], 2)
    ])

    workbook.save(excel_path)
    print(f"📁 Global comparison results saved to Excel: {excel_path}")


def save_comparison_figures(area_fig, perim_fig, main_output_path, case1, case2):
    """
    Saves area and perimeter comparison figures to Results/Figures/case1_vs_case2 folder.
    Args:
        area_fig (matplotlib.figure.Figure): Figure object for area comparison.
        perim_fig (matplotlib.figure.Figure): Figure object for perimeter comparison.
        main_output_path (str): Base output directory for the project (e.g., ~/Desktop/Project).
        case1 (str): Name of the first case (e.g., "OPEN").
        case2 (str): Name of the second case (e.g., "MISS").
    """
    # Ensure figures are valid matplotlib Figure objects
    if not hasattr(area_fig, "savefig") or not hasattr(perim_fig, "savefig"):
        raise TypeError("❌ Both area_fig and perim_fig must be matplotlib.figure.Figure objects.")

    # Prepare figure output folder
    folder_name = f"{case1}_vs_{case2}"
    fig_folder = os.path.join(main_output_path, "Results", "Figures", folder_name)
    os.makedirs(fig_folder, exist_ok=True)

    # Define file paths
    area_path = os.path.join(fig_folder, f"{folder_name}_area.png")
    perim_path = os.path.join(fig_folder, f"{folder_name}_perim.png")

    # Save figures
    area_fig.savefig(area_path, dpi=300, bbox_inches="tight")
    perim_fig.savefig(perim_path, dpi=300, bbox_inches="tight")

    print(f"🖼️ Saved comparison figures to: {fig_folder}")


def save_psoas_global_excel(stats_dict, excel_path):
    workbook = load_workbook(excel_path)

    sheet_name = "Psoas Atrophy Analysis"

    # Get or create Sheet 2
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)
        headers = [
            "Group", "N", "Mean", "Std Dev", "Median", "IQR", "", 
            "Mean Diff (I - C)", "Std Dev Diff", "Cliff's Delta"
        ]
        sheet.append(headers)

    # Leave Volume Analysis (sheet 1) untouched
    if sheet.max_row > 1:
        sheet.append([])  # separator row

    # Append stats
    sheet.append([
        "Intervened",
        stats_dict["intervened"]["n"],
        stats_dict["intervened"]["global_mean"],
        stats_dict["intervened"]["global_std"],
        stats_dict["intervened"]["median"],
        stats_dict["intervened"]["iqr"],
        "",
        stats_dict["diff"]["global_mean"],
        stats_dict["diff"]["global_std"],
        stats_dict["cliffs_delta"]
    ])

    sheet.append([
        "Control",
        stats_dict["control"]["n"],
        stats_dict["control"]["global_mean"],
        stats_dict["control"]["global_std"],
        stats_dict["control"]["median"],
        stats_dict["control"]["iqr"]
    ])

    workbook.save(excel_path)
    print(f"📁 Global psoas atrophy results appended to Excel sheet: '{sheet_name}'")



def save_psoas_figures(figures_dict, main_output_path):
    folder_name = "Intervened vs Control"
    fig_folder = os.path.join(main_output_path, "Figures", folder_name)
    os.makedirs(fig_folder, exist_ok=True)

    filenames = {
        "hist_intervened": "Histogram_Intervened.png",
        "hist_control": "Histogram_Control.png",
        "hist_diff": "Histogram_Difference.png",
        "boxplot": "Boxplot_Intervened_vs_Control.png"
    }

    for key, fig in figures_dict.items():
        if fig is not None:
            save_path = os.path.join(fig_folder, filenames.get(key, f"{key}.png"))
            fig.savefig(save_path)
            print(f"🖼️ Saved figure: {save_path}")

    print(f"🖼️ All psoas stat analysis figures saved to: {fig_folder}")