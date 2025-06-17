import os
import SimpleITK as sitk
import numpy as np
import tempfile
import json
from Patient_Analysis.ShapeAnalysis import rotate_segmentations, shape_statistics
from openpyxl import load_workbook
from statistics import mean, stdev
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def analysis_process(patient,excel_path):
    volume_value_from_patient(patient, excel_path)
    if patient.roi == "spinal_cord":
        spinal_cord_shape_analysis(patient)
    elif patient.roi == "iliopsoas":
        psoas_atrophy_analysis(patient)


def volume_value_from_patient(patient, excel_path):
    import os
    import numpy as np
    import SimpleITK as sitk
    from statistics import mean, stdev
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill

    case_type = patient.case_type
    patient_id = patient.patient_number
    crop_folders = patient.crop_paths

    if case_type == "LLIFR":
        intervened_psoas = "Right"
        case_type = "LLIF"
    elif case_type == "LLIFL":
        intervened_psoas = "Left"
        case_type = "LLIF"
    else:
        intervened_psoas = None  # for ALIF, PLIF, etc.

    if not crop_folders:
        print(f"⚠️ No crop folders found for Patient {patient_id}")
        return None

    if not os.path.exists(excel_path):
        print(f"❌ Excel file not found: {excel_path}")
        return None

    workbook = load_workbook(excel_path)
    sheet = workbook.active
    sheet.title = "Volume Analysis"

    # Add header if empty
    if sheet.max_row == 1 and not sheet.cell(1, 1).value:
        sheet.append(['CASE TYPE', 'PATIENT ID', 'LEVEL', 'PRE [cm³]', 'POST [cm³]', 'INCREMENT [%]'])

    for col_idx in range(1, 11):
        col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = 16

    results = []

    for crop_folder in crop_folders:
        if not os.path.exists(crop_folder):
            continue

        # Get anatomical level (e.g., L4-L5)
        level_base = os.path.basename(crop_folder).replace("Crop_", "").replace("_", "-")
        print(f"🔍 Processing crop folder: {crop_folder} for level {level_base}")

        # Determine name for this level row
        if patient.roi == "iliopsoas":
            side = "Left" if "Left" in crop_folder else "Right"
            if case_type == "LLIF" and intervened_psoas:
                level_name = f"{level_base}_Intervened_psoas" if side == intervened_psoas else f"{level_base}_Control_psoas"
            else:
                level_name = f"{level_base}_{side}_psoas"
        else:
            level_name = level_base

        # Get cropped files
        pre_path = os.path.join(crop_folder, "PRE_cropped.nrrd")
        post_path = os.path.join(crop_folder, "POST_cropped.nrrd")

        if not (os.path.exists(pre_path) and os.path.exists(post_path)):
            print(f"⚠️ Missing PRE or POST file in {crop_folder}")
            continue

        volumes = {}
        for label, path in [("PRE", pre_path), ("POST", post_path)]:
            image = sitk.ReadImage(path)
            spacing = image.GetSpacing()
            voxel_volume_cm3 = spacing[0] * spacing[1] * spacing[2] / 1000
            image_array = sitk.GetArrayFromImage(image)
            total_volume = np.sum(image_array) * voxel_volume_cm3
            volumes[label] = total_volume

        increment = (volumes["POST"] - volumes["PRE"]) / volumes["PRE"] * 100 if volumes["PRE"] != 0 else float('nan')
        pre_str = f"{volumes['PRE']:.5f}".replace('.', ',')
        post_str = f"{volumes['POST']:.5f}".replace('.', ',')
        increment_str = f"{increment:.2f}".replace('.', ',')

        # Update or append to Excel
        found = False
        for row in range(2, sheet.max_row + 1):
            if (sheet.cell(row, 1).value == case_type and
                sheet.cell(row, 2).value == f'{patient_id}' and
                sheet.cell(row, 3).value == level_name):
                sheet.cell(row, 4, pre_str)
                sheet.cell(row, 5, post_str)
                sheet.cell(row, 6, increment_str)
                found = True
                break

        if not found:
            sheet.append([case_type, f'{patient_id}', level_name, pre_str, post_str, increment_str])

        results.append((level_name, pre_str, post_str, increment_str))

    # Compute means/stdevs
    case_types = ['LLIF INDIRECT', 'LLIF DIRECT', 'LLIF', 'ALIF', 'PLIF', 'MIS', 'OPEN']
    increments_by_type = {t: [] for t in case_types}

    for row in range(2, sheet.max_row + 1):
        t = sheet.cell(row, 1).value
        val = sheet.cell(row, 6).value
        if t in increments_by_type and val:
            try:
                increments_by_type[t].append(float(val.replace(',', '.')))
            except ValueError:
                continue

    sheet.cell(1, 8, 'CASE TYPE')
    sheet.cell(1, 9, 'MEAN [%]')
    sheet.cell(1, 10, 'STD [%]')

    for i, t in enumerate(case_types, start=1):
        valid_vals = [v for v in increments_by_type[t] if not np.isnan(v)]
        mean_val = mean(valid_vals) if valid_vals else float('nan')
        std_val = stdev(valid_vals) if len(valid_vals) > 1 else float('nan')
        sheet.cell(i + 1, 8, t)
        sheet.cell(i + 1, 9, f"{mean_val:.2f}".replace('.', ','))
        sheet.cell(i + 1, 10, f"{std_val:.2f}".replace('.', ','))

    # Format
    bold_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = bold_font
    for cell in sheet['G']:
        cell.font = bold_font
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    workbook.save(excel_path)

    print(f"\n📊 Volume results for Patient {patient_id}:")
    for level, pre, post, inc in results:
        print(f"  Level {level}: PRE = {pre} cm³ | POST = {post} cm³ | Δ% = {inc}")

    return results



def spinal_cord_shape_analysis(patient):
    """
    Runs shape analysis for spinal cord ROI only.
    Assumes crop folders contain registered segmentations for each level.
    """
    patient_id = patient.patient_number
    case_type = patient.case_type
    crop_folders = patient.crop_paths
    stats_path = patient.output_path

    if not crop_folders:
        print(f"⚠️ No crop folders found for Patient {patient_id}")
        return None

    results = []

    for crop_folder in crop_folders:
        level_base = os.path.basename(crop_folder).replace("Crop_", "").replace("_", "-")
        level = level_base

        stats_dir = os.path.join(stats_path, f"Stat Analysis {level}")
        os.makedirs(stats_dir, exist_ok=True)

        with tempfile.TemporaryDirectory() as temp_dir:
            rotate_segmentations(crop_folder, temp_dir)
            interp_length, path_area, path_perim = shape_statistics(temp_dir, stats_dir, patient_id)

            results.append({
                "patient_id": patient_id,
                "level": level,
                "interp_length": interp_length,
                "path_area": path_area,
                "path_perim": path_perim
            })

    patient.log_analysis(stats_dir)
    return results


def psoas_atrophy_analysis(patient):
    """
    Computes and stores iliopsoas volumes (PRE, POST, INCREMENT)
    per level in individual JSON files, labeled as Control or Intervened.
    """
    patient_id = patient.patient_number
    case_type = patient.case_type
    crop_folders = patient.crop_paths
    stats_path = patient.output_path

    # Determine intervened side from case type
    if case_type == "LLIFR":
        intervened_psoas = "Right"
    elif case_type == "LLIFL":
        intervened_psoas = "Left"
    else:
        intervened_psoas = None

    if not crop_folders or not intervened_psoas:
        print(f"⚠️ Skipping Patient {patient_id} — invalid crop folders or case type")
        return None

    stats_dir = os.path.join(stats_path, "Stat Analysis Psoas")
    os.makedirs(stats_dir, exist_ok=True)

    level_json_paths = []

    for crop_folder in crop_folders:
        if not os.path.exists(crop_folder):
            continue

        # Level name
        level_base = os.path.basename(os.path.dirname(crop_folder)).replace("Crop_", "").replace("_", "-")
        side = "Left" if "Left" in crop_folder else "Right"
        psoas_type = "Intervened_psoas" if side == intervened_psoas else "Control_psoas"

        pre_path = os.path.join(crop_folder, "PRE_cropped.nrrd")
        post_path = os.path.join(crop_folder, "POST_cropped.nrrd")

        if not (os.path.exists(pre_path) and os.path.exists(post_path)):
            print(f"⚠️ Missing PRE or POST file in {crop_folder}")
            continue

        try:
            pre_image = sitk.ReadImage(pre_path)
            post_image = sitk.ReadImage(post_path)

            spacing = pre_image.GetSpacing()
            voxel_volume_cm3 = spacing[0] * spacing[1] * spacing[2] / 1000

            pre_array = sitk.GetArrayFromImage(pre_image)
            post_array = sitk.GetArrayFromImage(post_image)

            pre_vol = np.sum(pre_array) * voxel_volume_cm3
            post_vol = np.sum(post_array) * voxel_volume_cm3
            increment = ((post_vol - pre_vol) / pre_vol) * 100 if pre_vol != 0 else float('nan')

            result = {
                "PATIENT_ID": patient_id,
                "LEVEL": level_base,
                "TYPE": psoas_type,
                "PRE": round(pre_vol, 5),
                "POST": round(post_vol, 5),
                "INCREMENT": round(increment, 2)
            }

            output_path = os.path.join(stats_dir, f"{level_base}_{psoas_type}.json")
            with open(output_path, "w") as f:
                json.dump(result, f, indent=2)
            level_json_paths.append(output_path)

        except Exception as e:
            print(f"❌ Error processing volume in {crop_folder}: {e}")
            continue

    if level_json_paths:
        print(f"✅ Saved psoas stats for Patient {patient_id} in: {stats_dir}")
        patient.log_analysis(stats_dir)
        return level_json_paths
    else:
        print(f"⚠️ No valid psoas levels processed for Patient {patient_id}")
        return None
